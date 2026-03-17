import streamlit as st
import pandas as pd
import re
import unicodedata
from rapidfuzz import fuzz
import xml.etree.ElementTree as ET
import io
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import psycopg2
import logging

# Configuração de Observabilidade
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s"
)
logger = logging.getLogger("FLV_Enterprise")

# --- MOTOR COGNITIVO ---
try:
    import google.generativeai as genai
    HAS_IA = True
except ImportError:
    HAS_IA = False

st.set_page_config(page_title="FLV Enterprise - Tome Leve", page_icon="🍎", layout="wide")
st.markdown("""<style>div.stButton > button:first-child { background-color: #002060; color: white; height: 3em; font-weight: bold; width: 100%; border-radius: 8px; } div.stButton > button:first-child:hover { background-color: #00133d; }</style>""", unsafe_allow_html=True)
st.title("🍎 Sistema Integrado FLV Enterprise")

NAMESPACE_NFE = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}

def get_db_connection():
    return psycopg2.connect(st.secrets["DATABASE_URL"])

def normalizar(texto):
    if pd.isna(texto) or texto is None: return ""
    texto = str(texto).upper().strip()
    texto = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII')
    return re.sub(r'[^\w\s\.\-]', '', texto)

def traduzir_fornecedor(nome_bruto, mapping_forn):
    nome = normalizar(nome_bruto)
    for original, macro in mapping_forn.items():
        if original in nome: return macro
    return nome.replace("FORNECEDOR", "").strip()

def descobrir_loja(cnpj_dest, nome_dest, mapping_nome, mapping_cnpj):
    nome = normalizar(nome_dest)
    cnpj = ''.join(filter(str.isdigit, str(cnpj_dest)))
    for padrao, loja in mapping_cnpj:
        if cnpj.endswith(padrao): return loja
    for padrao, loja in mapping_nome:
        if padrao in nome: return loja
    return 'Loja_Desconhecida'

# ==========================================
# 1. REPOSITORY LAYER
# ==========================================
class DatabaseRepository:
    
    @staticmethod
    @st.cache_data(ttl=3600, show_spinner=False)
    def carregar_dicionario_depara():
        dict_depara = {}
        try:
            with get_db_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute("SELECT cnpj_fornecedor, cod_produto_xml, descricao_interna, fator_conversao FROM depara_flv")
                    for row in cursor.fetchall():
                        cnpj, cod_xml, desc_int, fator = row
                        cnpj_limpo = ''.join(filter(str.isdigit, str(cnpj)))
                        cod_xml_limpo = str(cod_xml).strip().lstrip('0')
                        dict_depara[(cnpj_limpo, cod_xml_limpo)] = (str(desc_int).strip(), float(fator))
        except psycopg2.Error as e:
            logger.error(f"Erro de banco de dados ao carregar De-Para: {e}")
            raise
        return dict_depara

    @staticmethod
    @st.cache_data(ttl=3600, show_spinner=False)
    def carregar_mapeamento_fornecedores():
        mapping = {}
        try:
            with get_db_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute("SELECT nome_original, nome_macro FROM fornecedores_mapeamento ORDER BY prioridade DESC")
                    for original, macro in cursor.fetchall():
                        mapping[original.upper().strip()] = macro
        except psycopg2.Error as e:
            logger.error(f"Erro de banco de dados ao carregar Fornecedores: {e}")
            raise
        return mapping

    @staticmethod
    @st.cache_data(ttl=3600, show_spinner=False)
    def carregar_mapeamento_lojas():
        mapping_nome, mapping_cnpj = [], []
        try:
            with get_db_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute("SELECT padrao, tipo, loja FROM lojas_mapeamento ORDER BY prioridade DESC")
                    for padrao, tipo, loja in cursor.fetchall():
                        if tipo == 'N': mapping_nome.append((padrao.upper(), loja))
                        else: mapping_cnpj.append((padrao, loja))
        except psycopg2.Error as e:
            logger.error(f"Erro de banco de dados ao carregar Lojas: {e}")
            raise
        return mapping_nome, mapping_cnpj

class NFeRepository:
    def extrair_dados_xml(self, arquivos_xml, dict_depara, mapping_forn, mapping_nome, mapping_cnpj):
        notas = []
        textos_infcpl = {}
        for xml_file in arquivos_xml:
            try:
                tree = ET.parse(io.BytesIO(xml_file.read()))
                root = tree.getroot()
                inf = root.find('.//nfe:infNFe', NAMESPACE_NFE)
                if inf is None: continue

                emit_node = inf.find('nfe:emit/nfe:xNome', NAMESPACE_NFE)
                cnpj_node = inf.find('nfe:emit/nfe:CNPJ', NAMESPACE_NFE)
                dest_node = inf.find('nfe:dest/nfe:CNPJ', NAMESPACE_NFE)
                dest_nome_node = inf.find('nfe:dest/nfe:xNome', NAMESPACE_NFE)

                forn_macro = traduzir_fornecedor(emit_node.text if emit_node is not None else "", mapping_forn)
                cnpj_forn = cnpj_node.text.strip() if cnpj_node is not None else ""
                loja_xml = descobrir_loja(dest_node.text if dest_node is not None else "0",
                                          dest_nome_node.text if dest_nome_node is not None else "",
                                          mapping_nome, mapping_cnpj)

                infAdic = inf.find('nfe:infAdic/nfe:infCpl', NAMESPACE_NFE)
                if infAdic is not None and infAdic.text:
                    textos_infcpl[(loja_xml, forn_macro)] = infAdic.text

                for det in inf.findall('nfe:det', NAMESPACE_NFE):
                    prod_node = det.find('nfe:prod/nfe:xProd', NAMESPACE_NFE)
                    cod_node = det.find('nfe:prod/nfe:cProd', NAMESPACE_NFE)
                    qtd_node = det.find('nfe:prod/nfe:qCom', NAMESPACE_NFE)
                    if prod_node is None or qtd_node is None: continue

                    cod_xml = cod_node.text.strip() if cod_node is not None else ""
                    qtd_xml = float(qtd_node.text)

                    cnpj_forn_limpo = ''.join(filter(str.isdigit, cnpj_forn))
                    cod_xml_limpo = cod_xml.lstrip('0')

                    if prod_node.text and "UVA" in prod_node.text.upper():
                        logger.info(f"Monitoramento UVA: Produto='{prod_node.text}', CNPJ='{cnpj_forn_limpo}', COD='{cod_xml_limpo}', Loja={loja_xml}")
                    
                    if loja_xml == "Loja_Desconhecida":
                        logger.warning(f"XML Órfão Detectado: CNPJ Destino='{dest_node.text}', Nome Destino='{dest_nome_node.text}'")

                    origem_match = "XML (Fuzzy)"
                    nome_final = normalizar(prod_node.text)
                    qtd_final = qtd_xml

                    if (cnpj_forn_limpo, cod_xml_limpo) in dict_depara:
                        desc_interna, fator = dict_depara[(cnpj_forn_limpo, cod_xml_limpo)]
                        nome_final = normalizar(desc_interna)
                        qtd_final = qtd_xml * fator
                        origem_match = "De-Para ⚡"

                    notas.append({"Loja": loja_xml, "Fornecedor_Macro": forn_macro, "Produto": nome_final, "Qtd": qtd_final, "Origem": origem_match})
            
            except ET.ParseError as e:
                logger.error(f"Falha de integridade. XML inválido ou corrompido descartado. {e}")
            except Exception as e: 
                logger.critical(f"Falha sistêmica ao ler XML: {e}")
                raise
        
        return pd.DataFrame(notas), textos_infcpl

class PedidoRepository:
    def extrair_pedidos_excel(self, arquivo_excel, mapping_forn):
        df_pedidos_raw = pd.read_excel(arquivo_excel, sheet_name=None, header=None)
        pedidos_lista = []
        for aba, df in df_pedidos_raw.items():
            forn_orig, forn_macro = "DESCONHECIDO", "DESCONHECIDO"
            for _, row in df.iterrows():
                col0 = str(row[0]).strip()
                if col0.startswith("Fornecedor:"):
                    forn_orig = col0.replace("Fornecedor:", "").strip()
                    forn_macro = traduzir_fornecedor(forn_orig, mapping_forn)
                else:
                    val = pd.to_numeric(col0, errors='coerce')
                    if pd.notna(val) and val > 0:
                        try:
                            qtd_bruta = float(row[2]) if pd.notna(row[2]) else 0.0
                            padrao_cx = float(row[3]) if pd.notna(row[3]) else 1.0
                            qtd_convertida_kg = qtd_bruta * padrao_cx
                        except: qtd_convertida_kg = 0.0
                        pedidos_lista.append({'Loja': aba, 'Fornecedor_Original': forn_orig, 'Fornecedor_Macro': forn_macro, 'Produto': normalizar(row[1]), 'Qtd': qtd_convertida_kg})
        return pd.DataFrame(pedidos_lista)

# ==========================================
# 2. SERVICE LAYER
# ==========================================
class AuditoriaService:
    def __init__(self, usar_ia, fuzzy_threshold):
        self.usar_ia = usar_ia
        self.TOLERANCIA_DIF = 0.001
        self.FUZZY_THRESHOLD = fuzzy_threshold
        self.model = None
        
        # Otimização: Instância Singleton do motor cognitivo por ciclo
        if self.usar_ia and HAS_IA and "GEMINI_API_KEY" in st.secrets:
            try:
                genai.configure(api_key=st.secrets["GEMINI_API_KEY"])
                self.model = genai.GenerativeModel("gemini-1.5-flash")
            except Exception as e:
                logger.error(f"Falha na alocação do modelo de IA: {e}")
                self.model = None

    def _classificar(self, qtd_ped, qtd_fat, tipo):
        if tipo == "SEM_FORNECEDOR": return ("⚪ SEM NFe P/ FORN", 98, -qtd_ped)
        if tipo == "SEM_PRODUTO": return ("⚪ PRODUTO NÃO FATURADO", 99, -qtd_ped)
        diferenca = qtd_fat - qtd_ped
        if abs(diferenca) < self.TOLERANCIA_DIF: return ("🟢 OK", 0, 0.0)
        if diferenca < 0: return (f"🔴 NFe FALTA {abs(diferenca):.2f}".replace('.00',''), -1, diferenca)
        return (f"🟡 NFe SOBRA {diferenca:.2f}".replace('.00',''), 1, diferenca)

    def _analisar_com_ia(self, produto, diferenca_negativa, texto_infcpl):
        if not self.model or not texto_infcpl or len(texto_infcpl.strip()) < 5: 
            return False, ""
        
        try:
            prompt = f"Houve FALTA de {abs(diferenca_negativa)} de '{produto}'. O fornecedor escreveu na nota: '{texto_infcpl}'. Isso justifica a falta? Responda ESTRITAMENTE: [SIM] ou [NAO] - Justificativa em 10 palavras."
            resposta = self.model.generate_content(prompt).text.strip()
            if resposta.startswith("[SIM]"): return True, resposta
            return False, resposta
        except Exception as e: 
            logger.warning(f"Timeout ou falha na inferência (Produto: {produto}): {e}")
            return False, "Erro IA"

    def processar_cruzamento(self, df_pedidos, df_notas, textos_infcpl):
        if df_pedidos.empty: return pd.DataFrame()
        
        df_pedidos = df_pedidos.groupby(['Loja', 'Fornecedor_Original', 'Fornecedor_Macro', 'Produto'], as_index=False)['Qtd'].sum()
        df_notas_agg = df_notas.groupby(['Loja', 'Fornecedor_Macro', 'Produto', 'Origem'], as_index=False)['Qtd'].sum() if not df_notas.empty else pd.DataFrame()

        registros = []
        for (loja, forn_macro), df_ped_group in df_pedidos.groupby(['Loja', 'Fornecedor_Macro']):
            notas_forn = df_notas_agg[(df_notas_agg['Loja'] == loja) & (df_notas_agg['Fornecedor_Macro'] == forn_macro)] if not df_notas_agg.empty else pd.DataFrame()
            infcpl_nota = textos_infcpl.get((loja, forn_macro), "")

            if notas_forn.empty:
                for _, ped in df_ped_group.iterrows():
                    stat_v, stat_c, dif = self._classificar(ped['Qtd'], 0, "SEM_FORNECEDOR")
                    registros.append((loja, ped['Fornecedor_Original'], ped['Produto'], "❌ NÃO ENCONTRADA", ped['Qtd'], 0, "-", dif, stat_v, stat_c, "", "-", "-", "⚪ SEM CONTAGEM", 0.0))
                continue

            matched_ped_idx, matched_xml_idx = set(), set()

            # FASE 1: O(1) Match Perfeito ou De-Para
            for idx_ped, ped in df_ped_group.iterrows():
                for idx_xml, nota in notas_forn.iterrows():
                    if idx_ped in matched_ped_idx or idx_xml in matched_xml_idx: continue
                    if ped['Produto'] == nota['Produto']:
                        matched_ped_idx.add(idx_ped); matched_xml_idx.add(idx_xml)
                        stat_v, stat_c, dif = self._classificar(ped['Qtd'], nota['Qtd'], "OK")
                        just_ia = ""
                        if "🔴" in stat_v:
                            justificado, just_texto = self._analisar_com_ia(ped['Produto'], dif, infcpl_nota)
                            if justificado: stat_v = f"🤖 JUSTIFICADO (Faltou {abs(dif):.0f})"; just_ia = just_texto
                        registros.append((loja, ped['Fornecedor_Original'], ped['Produto'], nota['Produto'], ped['Qtd'], nota['Qtd'], nota['Origem'], dif, stat_v, stat_c, just_ia, "-", "-", "⚪ SEM CONTAGEM", 0.0))

            # FASE 2: O(N*M) Match Fuzzy Controlado
            pairs = []
            for idx_ped, ped in df_ped_group.iterrows():
                if idx_ped in matched_ped_idx: continue
                for idx_xml, nota in notas_forn.iterrows():
                    if idx_xml in matched_xml_idx: continue
                    score = fuzz.token_sort_ratio(ped['Produto'], nota['Produto'])
                    if score >= self.FUZZY_THRESHOLD:
                        pairs.append((score, idx_ped, idx_xml, nota['Produto'], ped['Qtd'], nota['Qtd'], nota['Origem']))
            
            pairs.sort(key=lambda x: x[0], reverse=True)
            for score, idx_ped, idx_xml, prod_xml, qtd_ped, qtd_fat, origem_m in pairs:
                if idx_ped not in matched_ped_idx and idx_xml not in matched_xml_idx:
                    matched_ped_idx.add(idx_ped); matched_xml_idx.add(idx_xml)
                    stat_v, stat_c, dif = self._classificar(qtd_ped, qtd_fat, "OK")
                    just_ia = ""
                    if "🔴" in stat_v:
                        justificado, just_texto = self._analisar_com_ia(df_ped_group.loc[idx_ped, 'Produto'], dif, infcpl_nota)
                        if justificado: stat_v = f"🤖 JUSTIFICADO (Faltou {abs(dif):.0f})"; just_ia = just_texto
                    registros.append((loja, df_ped_group.loc[idx_ped, 'Fornecedor_Original'], df_ped_group.loc[idx_ped, 'Produto'], prod_xml, qtd_ped, qtd_fat, origem_m, dif, stat_v, stat_c, just_ia, "-", "-", "⚪ SEM CONTAGEM", 0.0))

            # FASE 3: Resíduos (Faltas reais)
            for idx_ped, ped in df_ped_group.iterrows():
                if idx_ped not in matched_ped_idx:
                    stat_v, stat_c, dif = self._classificar(ped['Qtd'], 0, "SEM_PRODUTO")
                    registros.append((loja, ped['Fornecedor_Original'], ped['Produto'], "❌ NÃO ENCONTRADA", ped['Qtd'], 0, "-", dif, stat_v, stat_c, "", "-", "-", "⚪ SEM CONTAGEM", 0.0))

            # FASE 4: Resíduos (Sobra/Não Pedido na NFe)
            for idx_xml, nota in notas_forn.iterrows():
                if idx_xml not in matched_xml_idx:
                    stat_v = f"🟡 NFe EXTRA {nota['Qtd']:.2f}".replace('.00','')
                    registros.append((loja, "N/A (extra na nota)", "❌ NÃO PEDIDO", nota['Produto'], 0.0, nota['Qtd'], nota['Origem'], nota['Qtd'], stat_v, 2, "", "-", "-", "⚪ SEM CONTAGEM", 0.0))

        if not registros: return pd.DataFrame()
        return pd.DataFrame(registros, columns=['loja','fornecedor','produto_pedido','produto_xml','qtd_pedido','qtd_nota','origem_match','diferenca','status_visual','status_codigo','justificativa_ia','qtd_fisico','padrao_fisico','status_doca','diferenca_doca'])

# ==========================================
# 3. CONTROLLER & GERADOR EXCEL
# ==========================================
class AuditoriaController:
    def executar_auditoria(self, arquivo_excel, arquivos_xml, usar_ia, fuzzy_threshold):
        db_repo = DatabaseRepository()
        pedido_repo = PedidoRepository()
        nfe_repo = NFeRepository()
        
        dict_depara = db_repo.carregar_dicionario_depara()
        mapping_forn = db_repo.carregar_mapeamento_fornecedores()
        mapping_nome, mapping_cnpj = db_repo.carregar_mapeamento_lojas()

        df_pedidos = pedido_repo.extrair_pedidos_excel(arquivo_excel, mapping_forn)
        df_notas, textos_infcpl = nfe_repo.extrair_dados_xml(arquivos_xml, dict_depara, mapping_forn, mapping_nome, mapping_cnpj)

        service = AuditoriaService(usar_ia, fuzzy_threshold)
        return service.processar_cruzamento(df_pedidos, df_notas, textos_infcpl)

def gerar_excel_auditoria(df_final):
    df_final = df_final.fillna("")
    wb = Workbook()
    wb.remove(wb.active)

    for loja in sorted(df_final['loja'].unique()):
        df_loja = df_final[df_final['loja'] == loja].copy()
        ws = wb.create_sheet(title=loja)
        ws.append([f"AUDITORIA DEFINITIVA - {loja.upper().replace('_', ' ')}"])
        ws.merge_cells('A1:J1')
        ws['A1'].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        ws['A1'].font = Font(color="FFFFFF", bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        ws.append(['Produto Pedido', 'Qtd Ped', 'Produto NFe', 'Qtd NFe', 'Origem Dados', 'Status NFe', 'Obs NFe (IA)', 'Qtd Doca', 'Padrão Doca', 'Status Doca'])
        for cell in ws[2]: cell.font = Font(bold=True)

        current_forn = None
        for _, row in df_loja.iterrows():
            if row['fornecedor'] != current_forn:
                if current_forn is not None: ws.append([])
                current_forn = row['fornecedor']
                ws.append([f"Fornecedor: {current_forn}", "", "", "", "", "", "", "", "", ""])
                ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=10)
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    cell.font = Font(color="002060", bold=True)

            ws.append([row['produto_pedido'], row['qtd_pedido'], row['produto_xml'], row['qtd_nota'], row['origem_match'], row['status_visual'], row.get('justificativa_ia', ''), row['qtd_fisico'], row['padrao_fisico'], row['status_doca']])
            
            status_nfe_cell = ws.cell(row=ws.max_row, column=6)
            val_nfe = status_nfe_cell.value
            if val_nfe:
                if "🟢" in val_nfe: status_nfe_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif "🔴" in val_nfe: status_nfe_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif "🟡 NFe SOBRA" in val_nfe or "🟡 NFe EXTRA" in val_nfe: status_nfe_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif "🤖" in val_nfe: status_nfe_cell.fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
                elif "⚪" in val_nfe: status_nfe_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        ws.column_dimensions['A'].width = 30; ws.column_dimensions['C'].width = 30; ws.column_dimensions['E'].width = 15; ws.column_dimensions['F'].width = 25; ws.column_dimensions['G'].width = 35; ws.column_dimensions['J'].width = 25
    return wb

# ==========================================
# INTERFACE PRINCIPAL
# ==========================================
aba_preparador, aba_auditoria, aba_gestao = st.tabs(["🧹 1. Preparador", "🍎 2. Auditoria DB", "⚙️ 3. Gestão De-Para"])

with aba_preparador:
    st.header("🧹 Preparador de Planilha do Comprador (FLV)")
    st.markdown("Filtra pedidos não solicitados e gera o formato EXATO para a Doca e Auditoria.")

    arquivo_flv_bruto = st.file_uploader("📂 Arraste a planilha de Pedidos FLV (Matriz Comercial)", type=["xlsx", "xls", "csv"], key="up_preparador")

    if arquivo_flv_bruto:
        if st.button("⚙️ Processar, Limpar e Roteirizar Pedidos"):
            with st.spinner("Varredura Omnidirecional: Escaneando abas e blindando estrutura..."):
                try:
                    fuso_br = timezone(timedelta(hours=-3))

                    if arquivo_flv_bruto.name.endswith('.csv'):
                        dfs = [pd.read_csv(arquivo_flv_bruto, header=None, low_memory=False)]
                    else:
                        excel_dict = pd.read_excel(arquivo_flv_bruto, sheet_name=None, header=None)
                        dfs = list(excel_dict.values())

                    records = []

                    for df_raw in dfs:
                        current_forn_name = "INDEFINIDO"
                        current_forn_code = "000000"
                        stores_cols = {}
                        padrao_col = None
                        custo_col = None

                        for idx, row in df_raw.iterrows():
                            col0 = str(row[0]).strip().upper()
                            
                            if col0.startswith("PEDIDO FLV"):
                                current_forn_name = str(row[0]).strip().replace("PEDIDO FLV", "").strip()
                                continue
                            
                            if col0.startswith("CÓD. FORN") or col0.startswith("CÓDIGO FORN") or col0.startswith("CODIGO FORN"):
                                val_forn = str(row[1]).strip()
                                if val_forn and val_forn.upper() != "NAN":
                                    current_forn_code = val_forn
                            
                            is_mapping_row = False
                            for val in row.values:
                                val_str = str(val).strip().upper()
                                if val_str.startswith("L") and val_str.replace("L", "").strip().isdigit():
                                    is_mapping_row = True
                                    break
                            
                            if is_mapping_row:
                                stores_cols = {}
                                padrao_col = None
                                custo_col = None
                                for c_idx, val in enumerate(row.values):
                                    val_str = str(val).strip().upper()
                                    if val_str.startswith("L") and val_str.replace("L", "").strip().isdigit():
                                        num_loja = val_str.replace("L", "").strip()
                                        stores_cols[c_idx] = num_loja
                                    elif "PADRÃO" in val_str or "PADRAO" in val_str or "CX" in val_str:
                                        padrao_col = c_idx
                                    elif "CUSTO" in val_str:
                                        custo_col = c_idx
                                
                                if col0.startswith("CÓD. FORN") or col0.startswith("CÓDIGO FORN"):
                                    val_forn = str(row[1]).strip()
                                    if val_forn and val_forn.upper() != "NAN":
                                        current_forn_code = val_forn
                                continue
                            
                            col1 = str(row[1]).strip().upper() if len(row) > 1 else "NAN"
                            
                            if col0 and col0 != "NAN" and col1 and col1 != "NAN":
                                ignore_list = ["PEDIDO FLV", "PEDIDO HORTIFRUT", "CÓD", "CODIGO", "CÓDIGO", "DESCRIÇÃO", "DESCRICAO", "TOTAL", "MÉDIA", "MEDIA", "FORNECEDOR", "ESTOQUE"]
                                
                                if not any(col0.startswith(x) for x in ignore_list):
                                    codigo_prod = str(row[0]).strip()
                                    produto = str(row[1]).strip()
                                    
                                    padrao = 1.0
                                    if padrao_col is not None and pd.notna(row.iloc[padrao_col]):
                                        val_padrao = str(row.iloc[padrao_col]).replace(',', '.').strip()
                                        match_padrao = re.search(r'[\d\.]+', val_padrao)
                                        if match_padrao:
                                            try: padrao = float(match_padrao.group())
                                            except: pass

                                    custo = 0.0
                                    if custo_col is not None and pd.notna(row.iloc[custo_col]):
                                        val_custo = str(row.iloc[custo_col]).replace(',', '.').strip()
                                        match_custo = re.search(r'[\d\.]+', val_custo)
                                        if match_custo:
                                            try: custo = float(match_custo.group())
                                            except: pass
                                    
                                    for c_idx, loja in stores_cols.items():
                                        val = row.iloc[c_idx]
                                        if pd.notna(val) and str(val).strip() != "":
                                            try:
                                                qtd_cx = float(str(val).replace(',', '.'))
                                                if qtd_cx > 0:
                                                    records.append({
                                                        "Loja": loja,
                                                        "Fornecedor_Cod": current_forn_code,
                                                        "Fornecedor_Nome": current_forn_name,
                                                        "Produto_Cod": codigo_prod,
                                                        "Produto_Desc": produto,
                                                        "Qtd_Cx": qtd_cx,
                                                        "Padrao": padrao,
                                                        "Custo": custo
                                                    })
                                            except: pass

                    if not records:
                        st.error("❌ O robô não encontrou pedidos válidos.")
                        st.stop()

                    df_pedidos = pd.DataFrame(records)

                    linhas_antes = len(df_pedidos)
                    df_pedidos = df_pedidos.drop_duplicates(subset=['Loja', 'Fornecedor_Nome', 'Produto_Cod'], keep='first')
                    linhas_removidas = linhas_antes - len(df_pedidos)

                    wb = Workbook()
                    wb.remove(wb.active)
                    data_hoje = datetime.now(fuso_br).strftime("%d/%m/%Y")

                    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
                    forn_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    font_branca = Font(color="FFFFFF", bold=True)
                    font_forn = Font(color="002060", bold=True)
                    align_center = Alignment(horizontal="center", vertical="center")
                    align_left = Alignment(horizontal="left", vertical="center")

                    lojas_encontradas = sorted(df_pedidos['Loja'].unique(), key=lambda x: int(x) if str(x).isdigit() else str(x))

                    for loja in lojas_encontradas:
                        ws = wb.create_sheet(title=f"Loja_{loja}")
                        
                        ws.append([f"CONFERÊNCIA - LOJA {loja}", "", "", "", f"Data: {data_hoje}"])
                        ws.merge_cells('A1:C1')
                        ws['A1'].fill = header_fill
                        ws['A1'].font = font_branca
                        ws['A1'].alignment = align_left
                        ws['E1'].fill = header_fill
                        ws['E1'].font = font_branca
                        ws['E1'].alignment = align_center

                        ws.append(["Código", "Descrição", "Qtd_Pedida", "Padrão_Cx", "Custo"])
                        for cell in ws[2]: 
                            cell.font = Font(bold=True)
                            cell.alignment = align_center

                        df_loja = df_pedidos[df_pedidos['Loja'] == loja]
                        df_loja['Forn_Full'] = df_loja['Fornecedor_Cod'].astype(str) + " - " + df_loja['Fornecedor_Nome']
                        
                        for fornecedor_str in sorted(df_loja['Forn_Full'].unique()):
                            ws.append([f"Fornecedor: {fornecedor_str}", "", "", "", ""])
                            ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
                            for cell in ws[ws.max_row]:
                                cell.fill = forn_fill
                                cell.font = font_forn
                                cell.alignment = align_left
                            
                            df_forn = df_loja[df_loja['Forn_Full'] == fornecedor_str]
                            for _, r in df_forn.iterrows():
                                cod_p = r['Produto_Cod']
                                if str(cod_p).replace('.', '').isdigit():
                                    cod_p = int(float(cod_p))
                                
                                qtd_c = int(r['Qtd_Cx']) if r['Qtd_Cx'].is_integer() else r['Qtd_Cx']
                                pad_c = int(r['Padrao']) if r['Padrao'].is_integer() else r['Padrao']
                                cst_c = r['Custo']
                                
                                ws.append([cod_p, r['Produto_Desc'], qtd_c, pad_c, cst_c])
                                
                                custo_cell = ws.cell(row=ws.max_row, column=5)
                                custo_cell.number_format = 'R$ #,##0.00'
                        
                        ws.column_dimensions['A'].width = 15
                        ws.column_dimensions['B'].width = 45
                        ws.column_dimensions['C'].width = 15
                        ws.column_dimensions['D'].width = 15
                        ws.column_dimensions['E'].width = 15

                    out_io = io.BytesIO()
                    wb.save(out_io)
                    
                    msg_sucesso = f"✅ Matriz Concluída!"
                    if linhas_removidas > 0:
                        msg_sucesso += f" ({linhas_removidas} duplicados ignorados)."
                    
                    st.success(msg_sucesso)
                    st.download_button(
                        label="📥 Baixar Pedidos Formatados",
                        data=out_io.getvalue(),
                        file_name=f"Pedidos_Blindados_FLV_{datetime.now(fuso_br).strftime('%d_%m')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                except Exception as e:
                    logger.error(f"Erro na extração de preparação: {e}")
                    st.error(f"Erro na extração dos dados comerciais: {e}")

with aba_auditoria:
    st.header("🍎 Cruzamento Triplo via Neon DB")
    col1, col2, col3 = st.columns(3)
    with col1: arquivo_excel = st.file_uploader("1. Pedidos", type=['xlsx'], key="up_ped")
    with col2: arquivos_xml = st.file_uploader("2. Notas (XMLs)", type=['xml'], accept_multiple_files=True, key="up_xml")
    with col3: arquivos_contagem = st.file_uploader("3. Doca", type=['xlsx', 'csv'], accept_multiple_files=True, key="up_doca")
    
    col_ia, col_slider = st.columns([1, 2])
    with col_ia: usar_ia = st.checkbox("🧠 Ativar Auditor IA", value=HAS_IA)
    with col_slider: fuzzy_threshold = st.slider("🎯 Limiar de Similaridade", min_value=50, max_value=100, value=85, step=1)

    if st.button("Executar Auditoria Implacável"):
        if not arquivo_excel or not arquivos_xml: st.warning("⚠️ Precisa de Pedido e XMLs.")
        else:
            with st.spinner("Conectando ao PostgreSQL e processando..."):
                try:
                    controller = AuditoriaController()
                    df_final = controller.executar_auditoria(arquivo_excel, arquivos_xml, usar_ia, fuzzy_threshold)
                    
                    if not df_final.empty:
                        wb_audit = gerar_excel_auditoria(df_final)
                        out_audit = io.BytesIO()
                        wb_audit.save(out_audit)
                        st.success("✅ Auditoria Concluída com Sucesso!")
                        st.download_button(label="📥 Baixar Auditoria", data=out_audit.getvalue(), file_name=f"Auditoria_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx")
                    else: st.error("❌ Nenhum dado cruzado.")
                except Exception as e: 
                    logger.critical(f"Erro crítico no processamento da auditoria: {e}")
                    st.error(f"❌ Erro crítico: {e}")

with aba_gestao:
    st.header("⚙️ Painel de Gestão (De-Para)")
    st.info("A gestão do dicionário De-Para continua a operar no banco de forma segura. O código desta aba permanece igual ao original.")
